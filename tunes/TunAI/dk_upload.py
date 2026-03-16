#!/usr/bin/env python3
"""
DistroKid Bulk Upload Automation
================================
Automates uploading singles to DistroKid using Playwright.

Artist: khayali / Liezl Claassen / 11138307 Records DK

Reads track metadata from DistroKid_Singles_Queue.xlsx ("DK Upload Queue" sheet)
OR falls back to dk_tracks.json.

Usage:
    1. pip install playwright openpyxl
    2. playwright install chromium
    3. python dk_upload.py                    # upload all pending tracks
    4. python dk_upload.py --dry-run          # preview what would be uploaded
    5. python dk_upload.py --track "Sideways" # upload a single track
    6. python dk_upload.py --start-from 5     # start from track index 5
    7. python dk_upload.py --source json      # force JSON instead of spreadsheet

The script will open a visible browser window. You'll need to log in to
DistroKid manually the first time — after that, the session cookie persists
in the user data directory.

Requirements:
    pip install playwright openpyxl
    playwright install chromium
"""

import json
import os
import sys
import time
import argparse
from pathlib import Path

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError:
    print("Playwright not installed. Run:")
    print("  pip install playwright")
    print("  playwright install chromium")
    sys.exit(1)


# ─── Configuration ───────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
CONFIG_FILE = SCRIPT_DIR / "dk_tracks.json"
XLSX_FILE = SCRIPT_DIR / "DistroKid_Singles_Queue.xlsx"
DKREADY_FOLDER = SCRIPT_DIR / "dkREADY"
USER_DATA_DIR = SCRIPT_DIR / ".dk_browser_profile"  # persists login session
DK_NEW_URL = "https://distrokid.com/new/"

# Delays (seconds) — tune these if DistroKid is slow/fast for you
DELAY_AFTER_NAVIGATE = 3
DELAY_AFTER_GENRE = 1
DELAY_AFTER_FILL = 1
DELAY_AFTER_FILE_UPLOAD = 5
DELAY_AFTER_ROBLOX = 2
DELAY_AFTER_SUBMIT = 10
DELAY_BETWEEN_TRACKS = 5

# ─── Genre value maps ────────────────────────────────────────────────────────

GENRE_NAME_TO_VAL = {
    "Alternative": "1", "Anime": "2", "Blues": "3", "Children's Music": "4",
    "Classical": "5", "Comedy": "6", "Country": "7", "Dance": "8",
    "Electronic": "9", "Folk": "11", "Hip-Hop/Rap": "12", "Holiday": "13",
    "Inspirational": "14", "Jazz": "15", "Latin": "16", "New Age": "17",
    "Opera": "18", "Pop": "25", "R&B/Soul": "26", "Reggae": "27",
    "Rock": "28", "Singer/Songwriter": "29", "Soundtrack": "30",
    "Spoken Word": "31", "World": "32",
}

SUBGENRE_NAME_TO_VAL = {
    "Big Room": "57", "Breaks": "35", "Chill Out": "36", "Deep House": "55",
    "Drum & Bass": "37", "Dubstep": "38", "Electro House": "39",
    "Electronica/Downtempo": "40", "Funk/Soul/Disco": "41", "Glitch Hop": "42",
    "Hard Dance": "43", "Hardcore/Hard Techno": "44", "Hip-Hop/R&B": "45",
    "House": "46", "Indie Dance/Nu Disco": "47", "Minimal/Deep Tech": "48",
    "Progressive House": "49", "Psy-Trance": "50", "Reggae/Dancehall/Dub": "51",
    "Tech House": "52", "Techno": "53", "Trance": "54",
}


# ─── Load track config ──────────────────────────────────────────────────────

def load_from_xlsx():
    """Load tracks from DistroKid_Singles_Queue.xlsx 'DK Upload Queue' sheet."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl not installed (pip install openpyxl). Falling back to JSON.")
        return None

    if not XLSX_FILE.exists():
        return None

    wb = load_workbook(XLSX_FILE, data_only=True)
    if 'DK Upload Queue' not in wb.sheetnames:
        print("No 'DK Upload Queue' sheet found in spreadsheet. Falling back to JSON.")
        return None

    ws = wb['DK Upload Queue']
    tracks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        title, genre_name, sub_name, instrumental, wav, png, wav_ok, png_ok, status, has_lyrics, isrc, notes = (
            row[:12] if len(row) >= 12 else list(row) + [None] * (12 - len(row))
        )
        if not title or not wav:
            continue

        genre_val = GENRE_NAME_TO_VAL.get(genre_name, "9")
        sub_val = SUBGENRE_NAME_TO_VAL.get(sub_name) if sub_name else None

        tracks.append({
            "title": str(title).strip(),
            "genre": genre_val,
            "subgenre": sub_val,
            "instrumental": str(instrumental).strip().lower() == 'yes' if instrumental else True,
            "wav": str(wav).strip(),
            "png": str(png).strip() if png else "",
            "status": str(status).strip().lower() if status else "pending",
            "notes": str(notes) if notes else "",
        })

    config = {
        "artist_name": "khayali",
        "songwriter_first": "Liezl",
        "songwriter_last": "Claassen",
        "apple_performer_role": "Synthesizer",
        "apple_performer_name": "khayali",
        "apple_producer_role": "Producer",
        "apple_producer_name": "khayali",
        "tracks": tracks,
        "_source": "xlsx",
    }
    print(f"  Loaded {len(tracks)} tracks from spreadsheet")
    return config


def save_status_to_xlsx(title, new_status):
    """Update a track's status in the spreadsheet."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(XLSX_FILE)
        ws = wb['DK Upload Queue']
        for row in ws.iter_rows(min_row=2):
            if row[0].value and str(row[0].value).strip() == title:
                row[8].value = new_status  # Column I = status
                break
        wb.save(XLSX_FILE)
    except Exception as e:
        print(f"    Warning: could not update spreadsheet status: {e}")


def load_config(source="auto"):
    if source in ("xlsx", "auto"):
        config = load_from_xlsx()
        if config:
            return config
        if source == "xlsx":
            print("ERROR: --source xlsx specified but spreadsheet not usable.")
            sys.exit(1)

    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        config = json.load(f)
    config["_source"] = "json"
    print(f"  Loaded {len(config['tracks'])} tracks from dk_tracks.json")
    return config


def save_config(config):
    """Save updated status back to whichever source we loaded from."""
    if config.get("_source") == "xlsx":
        return  # status saved per-track in save_status_to_xlsx
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=2, ensure_ascii=False)


# ─── Core automation ─────────────────────────────────────────────────────────

def wait_for_login(page):
    """If not logged in, wait for user to log in manually."""
    print("  Checking if logged in...")
    try:
        # If we see the upload form or dashboard, we're logged in
        page.wait_for_selector("#genrePrimary", timeout=10000)
        print("  ✓ Already logged in")
        return True
    except PWTimeout:
        pass

    # Check if we're on a login page
    if "signin" in page.url.lower() or "login" in page.url.lower():
        print("\n" + "=" * 60)
        print("  MANUAL STEP: Please log in to DistroKid in the browser.")
        print("  The script will continue automatically once you're logged in.")
        print("=" * 60 + "\n")
        # Wait up to 5 minutes for login
        try:
            page.wait_for_url("**/new/**", timeout=300000)
            time.sleep(DELAY_AFTER_NAVIGATE)
            return True
        except PWTimeout:
            print("  ✗ Login timeout (5 minutes). Exiting.")
            return False
    return True


def set_react_input(page, selector, value):
    """Set value on a React-controlled input using native setter."""
    page.evaluate(f"""
        (() => {{
            const el = document.querySelector('{selector}');
            if (!el) return;
            const nativeSet = Object.getOwnPropertyDescriptor(
                window.HTMLInputElement.prototype, 'value'
            ).set;
            nativeSet.call(el, {json.dumps(value)});
            el.dispatchEvent(new Event('input', {{bubbles: true}}));
            el.dispatchEvent(new Event('change', {{bubbles: true}}));
        }})()
    """)


def set_select(page, selector, value):
    """Set value on a <select> element and trigger change."""
    page.evaluate(f"""
        (() => {{
            const el = document.querySelector('{selector}');
            if (!el) return;
            el.value = {json.dumps(value)};
            el.dispatchEvent(new Event('change', {{bubbles: true}}));
        }})()
    """)


def tick_checkbox_by_id(page, element_id):
    """Tick a checkbox by ID if not already checked."""
    page.evaluate(f"""
        (() => {{
            const cb = document.getElementById('{element_id}');
            if (cb && !cb.checked) {{
                cb.click();
            }}
        }})()
    """)


def tick_all_service_checkboxes(page):
    """Tick ALL service/store checkboxes including Snapchat, MediaNet, Roblox."""
    # Standard extra services
    for cb_id in ['chksnap', 'chkbeats']:
        tick_checkbox_by_id(page, cb_id)

    # Roblox — needs special handling due to eligibility dialog
    roblox_checked = page.evaluate("""
        (() => {
            const cb = document.getElementById('chkroblox');
            if (cb && !cb.checked) {
                cb.click();
                return false;  // was not checked, just clicked it
            }
            return cb ? cb.checked : true;  // already checked or doesn't exist
        })()
    """)

    # Also tick any OTHER checkboxes that might appear in the services section
    # This future-proofs against new services DistroKid adds
    page.evaluate("""
        (() => {
            // Find the services/stores section and tick everything
            const allCheckboxes = document.querySelectorAll(
                'input[type="checkbox"][id^="chk"]'
            );
            allCheckboxes.forEach(cb => {
                if (!cb.checked) cb.click();
            });
        })()
    """)


def handle_roblox_dialog(page):
    """Handle the Roblox eligibility dialog that appears after ticking Roblox."""
    time.sleep(1)

    # Check if the Roblox dialog appeared
    dialog_visible = page.evaluate("""
        (() => {
            // Look for the Roblox eligibility modal
            const labels = Array.from(document.querySelectorAll('label'));
            const robloxLabel = labels.find(l =>
                l.textContent.includes('publishing rights')
            );
            return !!robloxLabel;
        })()
    """)

    if not dialog_visible:
        print("    (No Roblox dialog appeared)")
        return

    print("    Handling Roblox eligibility dialog...")

    # Click all 4 checkboxes in the dialog
    page.evaluate("""
        (() => {
            const labels = Array.from(document.querySelectorAll('label'));
            labels.forEach(label => {
                const text = label.textContent.toLowerCase();
                if (text.includes('publishing rights') ||
                    text.includes('pro') ||
                    text.includes('monetize') ||
                    text.includes('terms of service') ||
                    text.includes('terms of use')) {
                    const checkbox = label.querySelector('input[type="checkbox"]') ||
                                     label.previousElementSibling;
                    if (checkbox && !checkbox.checked) {
                        checkbox.click();
                    }
                }
            });
        })()
    """)

    time.sleep(0.5)

    # Click CONTINUE button in the dialog
    page.evaluate("""
        (() => {
            const buttons = Array.from(document.querySelectorAll('button, a'));
            const continueBtn = buttons.find(b =>
                b.textContent.trim().toUpperCase() === 'CONTINUE'
            );
            if (continueBtn) continueBtn.click();
        })()
    """)

    time.sleep(DELAY_AFTER_ROBLOX)
    print("    ✓ Roblox dialog handled")


def fill_track_form(page, track, config):
    """Fill in all form fields for a single track."""
    title = track["title"]
    genre = track["genre"]
    subgenre = track.get("subgenre")
    instrumental = track.get("instrumental", True)

    print(f"  Filling form for: {title}")

    # 1. Service checkboxes (ALL of them)
    print("    Ticking service checkboxes...")
    tick_all_service_checkboxes(page)
    time.sleep(1)

    # 2. Handle Roblox dialog if it appeared
    handle_roblox_dialog(page)

    # 3. Genre
    print(f"    Setting genre={genre}, subgenre={subgenre}")
    set_select(page, "#genrePrimary", genre)
    time.sleep(DELAY_AFTER_GENRE)

    # 4. Subgenre (only for genres that have them)
    if subgenre:
        set_select(page, "#subGenrePrimary", subgenre)
        time.sleep(0.5)

    # 5. Track title
    print(f"    Setting title: {title}")
    set_react_input(page, 'input[placeholder="Track 1 title"]', title)

    # 6. Songwriter
    print("    Setting songwriter: Liezl Claassen")
    set_react_input(page, 'input[name="songwriter_real_name_first1"]', config["songwriter_first"])
    set_react_input(page, 'input[name="songwriter_real_name_last1"]', config["songwriter_last"])

    # 7. Apple Music credits
    print("    Setting Apple Music credits...")
    set_select(page, "#track-1-performer-1-role", config["apple_performer_role"])
    set_react_input(page, "#track-1-performer-1-name", config["apple_performer_name"])
    set_select(page, "#track-1-producer-1-role", config["apple_producer_role"])
    set_react_input(page, "#track-1-producer-1-name", config["apple_producer_name"])

    # 8. Instrumental radio button
    if instrumental:
        print("    Setting instrumental=YES")
        page.evaluate("""
            (() => {
                const radios = Array.from(document.querySelectorAll('input[type="radio"]'));
                const instrRadio = radios.find(r =>
                    r.name && r.name.startsWith('instrumental_') && r.value === '1'
                );
                if (instrRadio) {
                    instrRadio.checked = true;
                    instrRadio.dispatchEvent(new Event('change', {bubbles: true}));
                }
            })()
        """)
    else:
        print("    Setting instrumental=NO (track has vocals)")
        page.evaluate("""
            (() => {
                const radios = Array.from(document.querySelectorAll('input[type="radio"]'));
                const notInstrRadio = radios.find(r =>
                    r.name && r.name.startsWith('instrumental_') && r.value === '0'
                );
                if (notInstrRadio) {
                    notInstrRadio.checked = true;
                    notInstrRadio.dispatchEvent(new Event('change', {bubbles: true}));
                }
            })()
        """)

    time.sleep(DELAY_AFTER_FILL)


def upload_files(page, track):
    """Upload cover art and audio file."""
    png_path = DKREADY_FOLDER / track["png"]
    wav_path = DKREADY_FOLDER / track["wav"]

    # Verify files exist
    if not png_path.exists():
        print(f"    ✗ Cover art not found: {png_path}")
        return False
    if not wav_path.exists():
        print(f"    ✗ Audio file not found: {wav_path}")
        return False

    print(f"    Uploading cover art: {track['png']}")
    file_inputs = page.query_selector_all('input[type="file"]')

    if len(file_inputs) < 2:
        print("    ✗ Could not find file input elements")
        return False

    # Cover art = first file input, audio = second
    file_inputs[0].set_input_files(str(png_path))
    time.sleep(2)

    print(f"    Uploading audio: {track['wav']}")
    file_inputs[1].set_input_files(str(wav_path))
    time.sleep(DELAY_AFTER_FILE_UPLOAD)

    print("    ✓ Files uploaded")
    return True


def tick_mandatory_checkboxes(page):
    """Tick all the mandatory confirmation checkboxes at the bottom of the form."""
    print("    Ticking mandatory checkboxes...")
    mandatory_ids = [
        'areyousureyoutube',
        'areyousurenonstandardscaps',
        'areyousurepromoservices',
        'areyousuresnap',
        'areyousurerecorded',
        'areyousureotherartist',
        'areyousuretandc',
        'areyousureinstrumental',
    ]
    for cb_id in mandatory_ids:
        page.evaluate(f"""
            (() => {{
                const cb = document.getElementById('{cb_id}');
                if (cb && !cb.checked) {{
                    cb.checked = true;
                    cb.dispatchEvent(new Event('change', {{bubbles: true}}));
                }}
            }})()
        """)

    # Also catch any other "areyousure" checkboxes that might exist
    page.evaluate("""
        (() => {
            document.querySelectorAll('input[type="checkbox"][id^="areyousure"]').forEach(cb => {
                if (!cb.checked) {
                    cb.checked = true;
                    cb.dispatchEvent(new Event('change', {bubbles: true}));
                }
            });
        })()
    """)
    print("    ✓ Mandatory checkboxes ticked")


def click_continue_button(page):
    """Click the main Continue/Submit button to submit the upload."""
    print("    Clicking Continue...")

    # Scroll to bottom first
    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(1)

    # Find and click the main Continue button (not the Roblox one)
    clicked = page.evaluate("""
        (() => {
            // Look for the main submit/continue button at the bottom
            const buttons = Array.from(document.querySelectorAll('button, input[type="submit"]'));
            const mainBtn = buttons.find(b => {
                const text = b.textContent.trim().toLowerCase();
                const val = (b.value || '').toLowerCase();
                return (text === 'continue' || val === 'continue') &&
                       b.offsetParent !== null;  // visible
            });
            if (mainBtn) {
                mainBtn.click();
                return true;
            }

            // Fallback: look for any prominent continue-style button
            const allLinks = Array.from(document.querySelectorAll('a'));
            const contLink = allLinks.find(a =>
                a.textContent.trim().toLowerCase() === 'continue' &&
                a.offsetParent !== null
            );
            if (contLink) {
                contLink.click();
                return true;
            }
            return false;
        })()
    """)

    if clicked:
        print("    ✓ Continue clicked")
    else:
        print("    ✗ Could not find Continue button — you may need to click it manually")

    return clicked


def wait_for_upload_complete(page, timeout=300):
    """Wait for the upload to complete (progress bars finish)."""
    print("    Waiting for upload to complete...")
    start = time.time()

    while time.time() - start < timeout:
        # Check if we've landed on the Mixea page or the "Hooray" page
        current_url = page.url.lower()
        if "mixea" in current_url or "hooray" in current_url or "done" in current_url:
            print("    ✓ Upload complete!")
            return True

        # Check page content for completion indicators
        done = page.evaluate("""
            (() => {
                const text = document.body.innerText.toLowerCase();
                return text.includes('hooray') ||
                       text.includes('your music is on its way') ||
                       text.includes('mixea') ||
                       text.includes('use my originals');
            })()
        """)
        if done:
            print("    ✓ Upload complete!")
            return True

        time.sleep(5)
        elapsed = int(time.time() - start)
        print(f"    ... still uploading ({elapsed}s)")

    print(f"    ✗ Upload timeout after {timeout}s")
    return False


def handle_mixea_page(page):
    """Handle the Mixea enhancement page — select 'Use my originals' and continue."""
    time.sleep(2)

    is_mixea = page.evaluate("""
        (() => {
            const text = document.body.innerText.toLowerCase();
            return text.includes('mixea') || text.includes('use my originals');
        })()
    """)

    if not is_mixea:
        print("    (No Mixea page detected)")
        return

    print("    Handling Mixea page — selecting 'Use my originals'...")

    # Click the "Use my originals" / free / R0 option
    page.evaluate("""
        (() => {
            // Look for radio buttons or options related to originals/free
            const labels = Array.from(document.querySelectorAll('label, div, span'));
            const origLabel = labels.find(el =>
                el.textContent.toLowerCase().includes('use my originals') ||
                el.textContent.toLowerCase().includes('no thanks')
            );
            if (origLabel) origLabel.click();

            // Also try clicking any R0/free radio
            const radios = Array.from(document.querySelectorAll('input[type="radio"]'));
            const freeRadio = radios.find(r => r.value === '0' || r.value === 'R0');
            if (freeRadio) {
                freeRadio.checked = true;
                freeRadio.click();
            }
        })()
    """)

    time.sleep(1)

    # Click Continue on the Mixea page
    page.evaluate("""
        (() => {
            const buttons = Array.from(document.querySelectorAll('button, a'));
            const btn = buttons.find(b =>
                b.textContent.trim().toLowerCase().includes('continue') &&
                b.offsetParent !== null
            );
            if (btn) btn.click();
        })()
    """)

    time.sleep(2)
    print("    ✓ Mixea handled")


def upload_single_track(page, track, config, index, total):
    """Complete upload flow for a single track."""
    title = track["title"]
    print(f"\n{'='*60}")
    print(f"  Track {index}/{total}: {title}")
    print(f"{'='*60}")

    if track.get("status") == "uploaded":
        print("  ⏭ Already uploaded, skipping")
        return True

    # 1. Navigate to upload page
    print("  Navigating to DistroKid upload page...")
    page.goto(DK_NEW_URL)
    time.sleep(DELAY_AFTER_NAVIGATE)

    # Make sure we're on the right page
    try:
        page.wait_for_selector("#genrePrimary", timeout=15000)
    except PWTimeout:
        # Might need login
        if not wait_for_login(page):
            return False
        page.goto(DK_NEW_URL)
        time.sleep(DELAY_AFTER_NAVIGATE)

    # 2. Fill the form
    fill_track_form(page, track, config)

    # 3. Upload files
    if not upload_files(page, track):
        print("  ✗ File upload failed, skipping track")
        return False

    # 4. Tick mandatory checkboxes
    tick_mandatory_checkboxes(page)

    # 5. Give user a moment to review (optional — remove for fully unattended)
    print("\n    ► Form filled. Review in browser if desired.")
    print("    ► Submitting in 5 seconds... (Ctrl+C to abort)")
    try:
        time.sleep(5)
    except KeyboardInterrupt:
        print("\n    Aborted by user. Track NOT submitted.")
        return False

    # 6. Click Continue to submit
    if not click_continue_button(page):
        print("  ✗ Could not submit, manual intervention needed")
        input("    Press Enter after you've submitted manually...")

    # 7. Wait for upload to complete
    if not wait_for_upload_complete(page):
        print("  ✗ Upload may not have completed")
        input("    Press Enter to continue to next track...")

    # 8. Handle Mixea page
    handle_mixea_page(page)

    # 9. Mark as uploaded
    track["status"] = "uploaded"
    if config.get("_source") == "xlsx":
        save_status_to_xlsx(title, "uploaded")
    save_config(config)
    print(f"\n  ✓ {title} — DONE")

    return True


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="DistroKid Bulk Upload Automation")
    parser.add_argument("--dry-run", action="store_true", help="Preview tracks without uploading")
    parser.add_argument("--track", type=str, help="Upload a single track by title (partial match)")
    parser.add_argument("--start-from", type=int, default=0, help="Start from track index (0-based)")
    parser.add_argument("--headless", action="store_true", help="Run headless (no visible browser)")
    parser.add_argument("--slow", type=int, default=0, help="Slow down actions by N ms (for debugging)")
    parser.add_argument("--source", choices=["auto", "xlsx", "json"], default="auto",
                        help="Data source: xlsx (spreadsheet), json, or auto (try xlsx first)")
    args = parser.parse_args()

    # Load config
    config = load_config(source=args.source)
    tracks = config["tracks"]

    # Filter tracks
    if args.track:
        tracks = [t for t in tracks if args.track.lower() in t["title"].lower()]
        if not tracks:
            print(f"No tracks matching '{args.track}'")
            return

    pending = [t for t in tracks if t.get("status") != "uploaded"]

    if args.start_from:
        pending = pending[args.start_from:]

    # Summary
    print(f"\n{'='*60}")
    print(f"  DistroKid Bulk Upload — khayali")
    print(f"{'='*60}")
    print(f"  Total tracks in config: {len(config['tracks'])}")
    print(f"  Already uploaded: {len([t for t in config['tracks'] if t.get('status') == 'uploaded'])}")
    print(f"  To upload now: {len(pending)}")
    print(f"  dkREADY folder: {DKREADY_FOLDER}")
    print()

    if args.dry_run:
        print("  DRY RUN — tracks that would be uploaded:\n")
        for i, t in enumerate(pending):
            png_exists = "✓" if (DKREADY_FOLDER / t["png"]).exists() else "✗ MISSING"
            wav_exists = "✓" if (DKREADY_FOLDER / t["wav"]).exists() else "✗ MISSING"
            vocal = "vocal" if not t.get("instrumental", True) else "instr"
            print(f"    {i+1:2d}. {t['title']}")
            print(f"        PNG: {t['png']} [{png_exists}]")
            print(f"        WAV: {t['wav']} [{wav_exists}]")
            print(f"        Genre: {t['genre']}/{t.get('subgenre', 'none')} | {vocal}")
            if t.get("notes"):
                print(f"        Note: {t['notes']}")
            print()
        return

    if not pending:
        print("  All tracks already uploaded! Nothing to do.")
        return

    # Verify files before starting
    print("  Verifying files exist...")
    missing = []
    for t in pending:
        if not (DKREADY_FOLDER / t["png"]).exists():
            missing.append(f"  PNG missing: {t['png']} (for '{t['title']}')")
        if not (DKREADY_FOLDER / t["wav"]).exists():
            missing.append(f"  WAV missing: {t['wav']} (for '{t['title']}')")

    if missing:
        print("\n  ⚠ Missing files:")
        for m in missing:
            print(f"    {m}")
        resp = input("\n  Continue anyway (missing tracks will be skipped)? [y/N]: ")
        if resp.lower() != 'y':
            return

    # Launch browser
    print("\n  Launching browser...")
    with sync_playwright() as p:
        browser = p.chromium.launch_persistent_context(
            user_data_dir=str(USER_DATA_DIR),
            headless=args.headless,
            slow_mo=args.slow,
            viewport={"width": 1280, "height": 900},
            args=["--disable-blink-features=AutomationControlled"],
        )

        page = browser.pages[0] if browser.pages else browser.new_page()

        # Navigate and check login
        page.goto(DK_NEW_URL)
        time.sleep(DELAY_AFTER_NAVIGATE)

        if not wait_for_login(page):
            browser.close()
            return

        # Upload each track
        total = len(pending)
        success = 0
        failed = []

        for i, track in enumerate(pending, 1):
            try:
                if upload_single_track(page, track, config, i, total):
                    success += 1
                else:
                    failed.append(track["title"])
            except Exception as e:
                print(f"\n  ✗ Error uploading {track['title']}: {e}")
                failed.append(track["title"])
                resp = input("  Continue to next track? [Y/n]: ")
                if resp.lower() == 'n':
                    break

            if i < total:
                print(f"\n  Cooling down {DELAY_BETWEEN_TRACKS}s before next track...")
                time.sleep(DELAY_BETWEEN_TRACKS)

        # Summary
        print(f"\n{'='*60}")
        print(f"  UPLOAD SESSION COMPLETE")
        print(f"{'='*60}")
        print(f"  Successful: {success}/{total}")
        if failed:
            print(f"  Failed: {', '.join(failed)}")
        print()

        browser.close()


if __name__ == "__main__":
    main()
